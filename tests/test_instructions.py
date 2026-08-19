from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from builder.constants import SheetNames
from builder.instructions import (
    ACTIVITY_LOG_FIELDS,
    CUSTOMER_DATABASE_FIELDS,
    INSTRUCTION_SECTIONS,
    build_instructions,
)


def test_instructions_builds_expected_content(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    worksheet = build_instructions(workbook)

    assert worksheet.title == SheetNames.instructions
    values = [cell.value for cell in worksheet["A"]]

    for section, _ in INSTRUCTION_SECTIONS:
        assert section in values
    for field in CUSTOMER_DATABASE_FIELDS:
        assert field in values
    for field in ACTIVITY_LOG_FIELDS:
        assert field in values

    output = tmp_path / "instructions.xlsx"
    workbook.save(output)
    reopened = load_workbook(output, data_only=False)
    assert SheetNames.instructions in reopened.sheetnames


def test_instructions_requires_workbook() -> None:
    try:
        build_instructions(object())  # type: ignore[arg-type]
    except TypeError:
        pass
    else:
        raise AssertionError("Expected TypeError was not raised.")


def test_instructions_rejects_duplicate_sheet() -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.create_sheet(SheetNames.instructions)

    try:
        build_instructions(workbook)
    except ValueError as exc:
        assert SheetNames.instructions in str(exc)
    else:
        raise AssertionError("Expected ValueError was not raised.")
