"""Functional tests for workbook utility functions."""

from openpyxl import Workbook

from builder.utils import (
    apply_headers,
    apply_print_settings,
    apply_title,
    auto_fit_columns,
    create_excel_table,
    freeze_header,
    set_active_cell,
)


def main() -> None:
    """Run workbook utility checks."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Test"

    header_row = apply_title(
        worksheet,
        "CRM Test Worksheet",
        "Utility validation",
    )

    headers = [
        "Customer ID",
        "Company Name",
        "Email Address",
    ]

    apply_headers(
        worksheet,
        headers,
        header_row,
    )

    worksheet.cell(
        row=header_row + 1,
        column=1,
        value="CRM-000001",
    )
    worksheet.cell(
        row=header_row + 1,
        column=2,
        value="Veridata Business Solutions",
    )
    worksheet.cell(
        row=header_row + 1,
        column=3,
        value="example@example.com",
    )

    data_end_row = header_row + 1

    auto_fit_columns(worksheet)

    freeze_header(
        worksheet,
        header_row,
    )

    create_excel_table(
        worksheet,
        "CustomerTable",
        header_row,
        data_end_row,
        1,
        len(headers),
    )

    apply_print_settings(
        worksheet,
        orientation="landscape",
        fit_to_width=1,
        repeat_header_row=header_row,
    )

    set_active_cell(
        worksheet,
        "A1",
    )

    assert worksheet.freeze_panes == f"A{header_row + 1}"

    assert worksheet.column_dimensions["A"].width >= 12
    assert worksheet.column_dimensions["B"].width >= 12
    assert worksheet.column_dimensions["C"].width >= 12

    assert "CustomerTable" in worksheet.tables

    assert worksheet.page_setup.orientation == "landscape"
    assert worksheet.page_setup.fitToWidth == 1

    assert worksheet.print_title_rows == f"$1:${header_row}"

    selection = worksheet.sheet_view.selection[0]
    assert selection.activeCell == "A1"
    assert selection.sqref == "A1"

    print("Complete utils.py functional test passed.")


if __name__ == "__main__":
    main()