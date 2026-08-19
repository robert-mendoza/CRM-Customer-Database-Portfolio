"""Unit and integration tests for the Activity Log builder."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from builder.activity_log import (
    ACTIVITY_LOG_HEADERS,
    ACTIVITY_LOG_TABLE_NAME,
    build_activity_log,
)
from builder.constants import SheetNames


def test_activity_log_builds_empty_table(tmp_path: Path) -> None:
    """Verify the Activity Log sheet and table contract."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    worksheet = build_activity_log(workbook)

    assert worksheet.title == SheetNames.activity_log
    assert [
        worksheet.cell(row=4, column=index).value
        for index in range(1, len(ACTIVITY_LOG_HEADERS) + 1)
    ] == list(ACTIVITY_LOG_HEADERS)
    assert list(worksheet.tables) == [ACTIVITY_LOG_TABLE_NAME]
    assert worksheet.tables[ACTIVITY_LOG_TABLE_NAME].ref == "A4:F4"
    assert worksheet.max_row == 4
    assert worksheet.max_column == 6
    assert worksheet.freeze_panes == "A5"

    output = tmp_path / "activity_log.xlsx"
    workbook.save(output)

    reopened = load_workbook(output, data_only=False)
    assert SheetNames.activity_log in reopened.sheetnames
    assert (
        reopened[SheetNames.activity_log]
        .tables[ACTIVITY_LOG_TABLE_NAME]
        .ref
        == "A4:F4"
    )


def test_activity_log_requires_openpyxl_workbook() -> None:
    """Verify invalid input is rejected."""
    try:
        build_activity_log(object())  # type: ignore[arg-type]
    except TypeError as exc:
        assert "Workbook" in str(exc)
    else:
        raise AssertionError("Expected TypeError was not raised.")


def test_activity_log_rejects_duplicate_sheet() -> None:
    """Verify duplicate worksheet creation is rejected."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.create_sheet(SheetNames.activity_log)

    try:
        build_activity_log(workbook)
    except ValueError as exc:
        assert SheetNames.activity_log in str(exc)
    else:
        raise AssertionError("Expected ValueError was not raised.")
