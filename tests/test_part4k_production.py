from __future__ import annotations

from openpyxl import load_workbook

from builder.activity_log import ACTIVITY_LOG_HEADERS, ACTIVITY_LOG_TABLE_NAME
from builder.constants import SHEETS
from builder.workbook_builder import build_workbook


def test_part4k_production(tmp_path) -> None:
    output = tmp_path / "part4k_production.xlsx"
    build_workbook(output)
    workbook = load_workbook(output, data_only=False)

    assert workbook.sheetnames == [
        SHEETS.cover,
        SHEETS.validation_lists,
        SHEETS.customer_database,
        SHEETS.dashboard,
        SHEETS.data_quality,
        SHEETS.activity_log,
        SHEETS.instructions,
    ]

    activity_log = workbook[SHEETS.activity_log]
    headers = [
        activity_log.cell(row=4, column=index).value
        for index in range(1, len(ACTIVITY_LOG_HEADERS) + 1)
    ]
    assert headers == list(ACTIVITY_LOG_HEADERS)
    assert ACTIVITY_LOG_TABLE_NAME in activity_log.tables
    assert activity_log.tables[ACTIVITY_LOG_TABLE_NAME].ref == "A4:F4"

    instructions = workbook[SHEETS.instructions]
    assert instructions["A1"].value == "CRM Workbook Instructions"

    database = workbook[SHEETS.customer_database]
    assert database.tables["CustomerDatabaseTable"].ref == "A4:J4"
    assert len(database.data_validations.dataValidation) == 5
    assert len(workbook.defined_names) == 5
