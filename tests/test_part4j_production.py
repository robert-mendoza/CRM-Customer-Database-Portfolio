"""Production integration test for Part 4J data quality reporting."""

from openpyxl import load_workbook

from builder.workbook_builder import build_workbook


def test_part4j_production(tmp_path) -> None:
    output = tmp_path / "part4j_production.xlsx"
    build_workbook(output)

    workbook = load_workbook(output, data_only=False)

    assert workbook.sheetnames == [
        "01_Cover",
        "04_Validation_Lists",
        "03_Customer_Database",
        "02_Dashboard",
        "05_Data_Quality_Report",
        "06_Activity_Log",
        "07_Instructions",
    ]

    report = workbook["05_Data_Quality_Report"]
    assert report["A1"].value == "Data Quality Report"
    assert report["A4"].value == "Total Records"
    assert report["B4"].value.startswith("=COUNTIF")
    assert report["A5"].value == "Data Quality Status"
    assert report["B5"].value.startswith("=IF")

    expected_checks = [
        "Missing Customer IDs",
        "Missing First Names",
        "Missing Last Names",
        "Missing Email Addresses",
        "Missing Phone Numbers",
        "Missing Industries",
        "Missing Customer Types",
        "Missing Customer Statuses",
        "Missing Lead Sources",
        "Missing Priorities",
    ]
    assert [report.cell(row=row, column=1).value for row in range(7, 17)] == expected_checks

    database = workbook["03_Customer_Database"]
    assert database.tables["CustomerDatabaseTable"].ref == "A4:J4"
    assert len(database.data_validations.dataValidation) == 5

    assert len(workbook.defined_names) == 5
