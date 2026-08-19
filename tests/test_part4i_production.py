"""Production integration test for Part 4I dashboard integration."""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from builder.constants import SHEETS, WORKBOOK
from builder.customer_database import CUSTOMER_HEADERS, CUSTOMER_TABLE_NAME
from builder.validation_names import VALIDATION_DEFINED_NAMES
from builder.workbook_builder import build_workbook

OUTPUT = PROJECT_ROOT / "output" / "part4i_production.xlsx"


def test_part4i_production_workbook() -> None:
    """Build, reopen, and validate the complete production workbook."""
    output_path = build_workbook(OUTPUT)
    assert output_path.exists()

    workbook = load_workbook(output_path, data_only=False)

    assert workbook.sheetnames == [
        SHEETS.cover,
        SHEETS.validation_lists,
        SHEETS.customer_database,
        SHEETS.dashboard,
        SHEETS.data_quality,
        SHEETS.activity_log,
        SHEETS.instructions,
    ]

    cover = workbook[SHEETS.cover]
    assert cover["A1"].value == WORKBOOK.title
    assert cover["A2"].value == WORKBOOK.company
    assert cover["B4"].value == WORKBOOK.version
    assert cover["B5"].value == WORKBOOK.author
    assert cover["B6"].value == WORKBOOK.filename

    database = workbook[SHEETS.customer_database]
    assert tuple(
        database.cell(4, index).value
        for index in range(1, len(CUSTOMER_HEADERS) + 1)
    ) == CUSTOMER_HEADERS
    assert CUSTOMER_TABLE_NAME in database.tables
    assert database.tables[CUSTOMER_TABLE_NAME].ref == "A4:J4"
    assert len(database.data_validations.dataValidation) == 5

    assert len(VALIDATION_DEFINED_NAMES) == 5
    for name, reference in VALIDATION_DEFINED_NAMES:
        assert workbook.defined_names[name].attr_text == reference

    dashboard = workbook[SHEETS.dashboard]
    assert dashboard["A1"].value == "CRM Dashboard"
    assert dashboard["A2"].value == WORKBOOK.company
    assert dashboard["A4"].value == "KPI"
    assert dashboard["B4"].value == "Value"
    assert dashboard["A5"].value == "Total Customers"
    assert dashboard["A6"].value == "Active Customers"
    assert dashboard["A7"].value == "Prospect Customers"
    assert dashboard["A8"].value == "Inactive Customers"
    assert dashboard["B5"].value == (
        "=COUNTA('03_Customer_Database'!A5:A1004)"
    )
    assert dashboard["B6"].value == (
        '=COUNTIF(\'03_Customer_Database\'!H5:H1004,"Active")'
    )
    assert dashboard["B7"].value == (
        '=COUNTIF(\'03_Customer_Database\'!H5:H1004,"Prospect")'
    )
    assert dashboard["B8"].value == (
        '=COUNTIF(\'03_Customer_Database\'!H5:H1004,"Inactive")'
    )
