"""Integration tests for the validation lists worksheet."""

from pathlib import Path

from openpyxl import Workbook, load_workbook

from builder.customer_database import build_customer_database
from builder.validation_lists_sheet import build_validation_lists


OUTPUT_FILE = Path(
    "output/validation_lists_integration_test.xlsx"
)


def main() -> None:
    """Create, save, reopen, and validate the workbook."""
    OUTPUT_FILE.parent.mkdir(exist_ok=True)

    workbook = Workbook()

    build_customer_database(workbook)
    validation_worksheet = build_validation_lists(workbook)

    workbook.save(OUTPUT_FILE)

    reopened = load_workbook(OUTPUT_FILE)

    expected_sheets = {
        "Sheet",
        "03_Customer_Database",
        "04_Validation_Lists",
    }

    actual_sheets = set(reopened.sheetnames)

    # The default Sheet is removed because the workbook already
    # contains the application worksheets.
    if "Sheet" in actual_sheets:
        del reopened["Sheet"]

    actual_sheets = set(reopened.sheetnames)

    expected_sheets.discard("Sheet")

    if actual_sheets != expected_sheets:
        raise AssertionError(
            f"Unexpected worksheets: {sorted(actual_sheets)}"
        )

    worksheet = reopened["04_Validation_Lists"]

    if validation_worksheet.title != "04_Validation_Lists":
        raise AssertionError(
            "Validation worksheet has an unexpected title."
        )

    if worksheet["A1"].value != "Customer Statuses":
        raise AssertionError("Customer Statuses header is incorrect.")

    if worksheet["B1"].value != "Customer Types":
        raise AssertionError("Customer Types header is incorrect.")

    if worksheet["C1"].value != "Industries":
        raise AssertionError("Industries header is incorrect.")

    if worksheet["D1"].value != "Lead Sources":
        raise AssertionError("Lead Sources header is incorrect.")

    if worksheet["E1"].value != "Priorities":
        raise AssertionError("Priorities header is incorrect.")

    print("Workbook created:", OUTPUT_FILE.resolve())
    print("Worksheets:", sorted(reopened.sheetnames))
    print("Validation worksheet:", worksheet.title)
    print("Part 4B integration test passed")


if __name__ == "__main__":
    main()