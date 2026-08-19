from __future__ import annotations

from openpyxl import load_workbook

from builder.constants import SHEETS
from builder.workbook_builder import build_workbook


def test_part4l_production(tmp_path) -> None:
    """Verify 07_Instructions in the production workbook."""
    output = tmp_path / "part4l_production.xlsx"
    build_workbook(output)

    workbook = load_workbook(output, data_only=False)

    assert SHEETS.instructions in workbook.sheetnames
    assert workbook.sheetnames[-1] == SHEETS.instructions

    instructions = workbook[SHEETS.instructions]
    assert instructions["A1"].value == "CRM Workbook Instructions"

    values = [cell.value for cell in instructions["A"]]
    expected_sections = [
        "Workbook Overview",
        "Customer Database",
        "Data Validation",
        "Activity Log",
        "Dashboard",
        "Data Quality Report",
        "General Guidelines",
        "Customer Database Fields",
        "Activity Log Fields",
    ]

    for section in expected_sections:
        assert section in values

    assert workbook[SHEETS.customer_database].tables[
        "CustomerDatabaseTable"
    ].ref == "A4:J4"
    assert len(workbook[SHEETS.customer_database].data_validations.dataValidation) == 5
    assert len(workbook.defined_names) == 5
