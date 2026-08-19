"""Integration tests for validation defined names."""

from pathlib import Path

from openpyxl import Workbook, load_workbook

from builder.validation_lists_sheet import build_validation_lists
from builder.validation_names import (
    VALIDATION_DEFINED_NAMES,
    build_validation_defined_names,
)


OUTPUT_FILE = Path(
    "output/validation_names_integration_test.xlsx"
)


def main() -> None:
    """Create, save, reopen, and verify defined names."""
    OUTPUT_FILE.parent.mkdir(exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)

    build_validation_lists(workbook)
    build_validation_defined_names(workbook)

    workbook.save(OUTPUT_FILE)

    reopened = load_workbook(OUTPUT_FILE)

    for name, expected_reference in VALIDATION_DEFINED_NAMES:
        defined_name = reopened.defined_names[name]

        if defined_name.attr_text != expected_reference:
            raise AssertionError(
                f"{name}: unexpected reference "
                f"{defined_name.attr_text!r}"
            )

        print(f"{name}: {defined_name.attr_text}")

    print("Workbook created:", OUTPUT_FILE.resolve())
    print("Defined names:", len(VALIDATION_DEFINED_NAMES))
    print("Part 4D integration test passed")


if __name__ == "__main__":
    main()