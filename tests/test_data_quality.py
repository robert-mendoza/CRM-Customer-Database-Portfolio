"""Unit and integration tests for the data quality report builder."""

from openpyxl import Workbook, load_workbook

from builder.customer_database import build_customer_database
from builder.data_quality import build_data_quality_report
from builder.validation_lists_sheet import build_validation_lists


def _build_base_workbook() -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    build_validation_lists(workbook)
    build_customer_database(workbook)
    return workbook


def test_data_quality_handles_empty_template() -> None:
    workbook = _build_base_workbook()
    worksheet = build_data_quality_report(workbook)

    assert worksheet.title == "05_Data_Quality_Report"
    assert worksheet["A4"].value == "Total Records"
    assert worksheet["B4"].value is not None
    assert worksheet["B5"].value is not None
    assert worksheet["A7"].value == "Missing Customer IDs"


def test_data_quality_counts_missing_values(tmp_path) -> None:
    workbook = _build_base_workbook()
    database = workbook["03_Customer_Database"]

    database.append(["CUST-0001", "Robert", "Mendoza", "", "555-1000", "Software", "Individual", "Active", "Website", "High"])
    database.append(["CUST-0002", "", "Santos", "santos@example.com", "", "Finance", "Enterprise", "Prospect", "Referral", "Medium"])

    worksheet = build_data_quality_report(workbook)
    output = tmp_path / "data_quality.xlsx"
    workbook.save(output)

    reopened = load_workbook(output, data_only=False)
    report = reopened["05_Data_Quality_Report"]

    assert report["B4"].value.startswith("=COUNTIF")
    assert report["B7"].value is not None
    assert report["B8"].value is not None
    assert report["B9"].value is not None
