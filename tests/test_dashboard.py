"""Integration tests for the CRM dashboard builder."""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from builder.constants import SHEETS
from builder.customer_database import CUSTOMER_HEADERS, build_customer_database
from builder.dashboard import build_dashboard
from builder.validation_lists_sheet import build_validation_lists


def test_dashboard_builds_dynamic_kpis(tmp_path: Path) -> None:
    """Verify dashboard KPIs reference the customer database dynamically."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    build_validation_lists(workbook)
    database = build_customer_database(workbook, CUSTOMER_HEADERS)

    database.append([
        "CUST-0001", "Ana", "Santos", "ana@example.com", "123",
        "Finance", "Individual", "Active", "Website", "High",
    ])
    database.append([
        "CUST-0002", "Ben", "Cruz", "ben@example.com", "456",
        "Retail", "Small Business", "Prospect", "Referral", "Medium",
    ])
    database.append([
        "CUST-0003", "Cara", "Reyes", "cara@example.com", "789",
        "Healthcare", "Enterprise", "Inactive", "Email", "Low",
    ])
    database.append([
        "CUST-0004", "Dan", "Lee", "dan@example.com", "000",
        "Software", "Medium Business", "Active", "Phone", "Urgent",
    ])

    dashboard = build_dashboard(workbook)

    assert dashboard.title == SHEETS.dashboard
    assert dashboard["A4"].value == "KPI"
    assert dashboard["B4"].value == "Value"
    assert dashboard["A5"].value == "Total Customers"
    assert dashboard["A6"].value == "Active Customers"
    assert dashboard["A7"].value == "Prospect Customers"
    assert dashboard["A8"].value == "Inactive Customers"

    assert dashboard["B5"].value == (
        "=COUNTA('03_Customer_Database'!A5:A1004)"
    )
    assert dashboard["B6"].value == (
        '=COUNTIF(\'03_Customer_Database\'!H5:H1004,"Active")'
    )
    assert dashboard["B7"].value == (
        '=COUNTIF(\'03_Customer_Database\'!H5:H1004,"Prospect")'
    )
    assert dashboard["B8"].value == (
        '=COUNTIF(\'03_Customer_Database\'!H5:H1004,"Inactive")'
    )

    output = tmp_path / "dashboard.xlsx"
    workbook.save(output)

    reopened = load_workbook(output, data_only=False)
    assert SHEETS.dashboard in reopened.sheetnames
    assert reopened[SHEETS.dashboard]["B5"].value == dashboard["B5"].value


def test_dashboard_requires_customer_database() -> None:
    """Verify dashboard creation fails when its required sheet is absent."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    try:
        build_dashboard(workbook)
    except ValueError as exc:
        assert SHEETS.customer_database in str(exc)
    else:
        raise AssertionError("Expected ValueError was not raised.")
