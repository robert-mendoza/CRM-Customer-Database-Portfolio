"""
Workbook style definitions.

This module centralizes all reusable Excel styles used throughout
the CRM Customer Database Builder workbook.

Project:
    CRM Customer Database Builder

Author:
    Robert Mendoza
"""

from __future__ import annotations

from dataclasses import dataclass

from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)


# ---------------------------------------------------------------------
# Font Definitions
# ---------------------------------------------------------------------


@dataclass(frozen=True)
class FontStyles:
    """Workbook font definitions."""

    title: Font
    header: Font
    body: Font


# ---------------------------------------------------------------------
# Fill Definitions
# ---------------------------------------------------------------------


@dataclass(frozen=True)
class FillStyles:
    """Workbook fill definitions."""

    header: PatternFill


# ---------------------------------------------------------------------
# Border Definitions
# ---------------------------------------------------------------------


@dataclass(frozen=True)
class BorderStyles:
    """Workbook border definitions."""

    thin: Border


# ---------------------------------------------------------------------
# Alignment Definitions
# ---------------------------------------------------------------------


@dataclass(frozen=True)
class AlignmentStyles:
    """Workbook alignment definitions."""

    left: Alignment
    center: Alignment

# ---------------------------------------------------------------------
# Standard Workbook Styles
# ---------------------------------------------------------------------

_DEFAULT_FONT_NAME = "Arial"

_TITLE_FONT_SIZE = 16
_HEADER_FONT_SIZE = 11
_BODY_FONT_SIZE = 10

_HEADER_FILL_COLOR = "D9EAF7"

_BORDER_COLOR = "B7B7B7"


# ---------------------------------------------------------------------
# Style Registry
# ---------------------------------------------------------------------


@dataclass(frozen=True)
class WorkbookStyles:
    """Central registry for reusable workbook styles."""

    fonts: FontStyles
    fills: FillStyles
    borders: BorderStyles
    alignments: AlignmentStyles


def _build_styles() -> WorkbookStyles:
    """Create the standard workbook style registry.

    Returns:
        Configured workbook styles.
    """

    thin_side = Side(
        style="thin",
        color=_BORDER_COLOR,
    )

    return WorkbookStyles(
        fonts=FontStyles(
            title=Font(
                name=_DEFAULT_FONT_NAME,
                size=_TITLE_FONT_SIZE,
                bold=True,
            ),
            header=Font(
                name=_DEFAULT_FONT_NAME,
                size=_HEADER_FONT_SIZE,
                bold=True,
            ),
            body=Font(
                name=_DEFAULT_FONT_NAME,
                size=_BODY_FONT_SIZE,
            ),
        ),
        fills=FillStyles(
            header=PatternFill(
                fill_type="solid",
                fgColor=_HEADER_FILL_COLOR,
            ),
        ),
        borders=BorderStyles(
            thin=Border(
                left=thin_side,
                right=thin_side,
                top=thin_side,
                bottom=thin_side,
            ),
        ),
        alignments=AlignmentStyles(
            left=Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True,
            ),
            center=Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            ),
        ),
    )


styles = _build_styles()


__all__ = [
    "AlignmentStyles",
    "BorderStyles",
    "FillStyles",
    "FontStyles",
    "WorkbookStyles",
    "styles",
]