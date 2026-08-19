"""Data quality report worksheet builder.

This module creates the data quality report for the CRM Customer Database
Builder. The report evaluates populated customer records in the current
customer database schema and does not treat unused template rows as data.

Project:
    CRM Customer Database Builder

Author:
    Robert Mendoza
"""

from __future__ import annotations

from copy import copy

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from builder.constants import SheetNames
from builder.styles import styles
from builder.utils import apply_title


_DATA_START_ROW = 5
_DATA_END_ROW = 1004

_CHECKS: tuple[tuple[str, str], ...] = (
    ("Missing Customer IDs", "A"),
    ("Missing First Names", "B"),
    ("Missing Last Names", "C"),
    ("Missing Email Addresses", "D"),
    ("Missing Phone Numbers", "E"),
    ("Missing Industries", "F"),
    ("Missing Customer Types", "G"),
    ("Missing Customer Statuses", "H"),
    ("Missing Lead Sources", "I"),
    ("Missing Priorities", "J"),
)



def _missing_formula(column: str) -> str:
    """Return an Excel formula counting missing values in populated records."""
    return (
        f'=SUMPRODUCT(--(\'{SheetNames.customer_database}\'!$A${_DATA_START_ROW}:'
        f'$A${_DATA_END_ROW}<>""),--(\'{SheetNames.customer_database}\'!'
        f'${column}${_DATA_START_ROW}:${column}${_DATA_END_ROW}=""))'
    )



def build_data_quality_report(workbook: Workbook) -> Worksheet:
    """Build the data quality report worksheet.

    The report uses Customer ID as the record-presence indicator. Blank
    template rows therefore do not count as records or as missing data.

    Args:
        workbook: Target workbook containing the customer database sheet.

    Returns:
        The completed data quality report worksheet.

    Raises:
        ValueError: If the customer database sheet does not exist.
    """
    database_name = SheetNames.customer_database
    if database_name not in workbook.sheetnames:
        raise ValueError(
            f"Required worksheet '{database_name}' does not exist."
        )

    worksheet = workbook.create_sheet(title=SheetNames.data_quality)
    apply_title(worksheet, "Data Quality Report")

    worksheet["A3"] = "Metric"
    worksheet["B3"] = "Value"
    worksheet["A3"].font = styles.fonts.header
    worksheet["B3"].font = styles.fonts.header
    worksheet["A3"].fill = styles.fills.header
    worksheet["B3"].fill = styles.fills.header
    worksheet["A3"].border = styles.borders.thin
    worksheet["B3"].border = styles.borders.thin

    worksheet["A4"] = "Total Records"
    worksheet["B4"] = (
        f'=COUNTIF(\'{SheetNames.customer_database}\'!$A${_DATA_START_ROW}:'
        f'$A${_DATA_END_ROW},"<>")'
    )

    worksheet["A5"] = "Data Quality Status"
    worksheet["B5"] = (
        '=IF(B4=0,"No Data",'
        'IF(SUM(B7:B16)=0,"Good","Needs Review"))'
    )

    worksheet["A6"] = "Check"
    worksheet["B6"] = "Missing Values"
    worksheet["A6"].font = styles.fonts.header
    worksheet["B6"].font = styles.fonts.header
    worksheet["A6"].fill = styles.fills.header
    worksheet["B6"].fill = styles.fills.header
    worksheet["A6"].border = styles.borders.thin
    worksheet["B6"].border = styles.borders.thin

    for row, (label, column) in enumerate(_CHECKS, start=7):
        worksheet.cell(row=row, column=1, value=label)
        worksheet.cell(row=row, column=2, value=_missing_formula(column))

    for row in worksheet.iter_rows(min_row=4, max_row=16, min_col=1, max_col=2):
        for cell in row:
            font = copy(styles.fonts.body)
            cell.font = font
            cell.border = styles.borders.thin
            cell.alignment = styles.alignments.left

    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions["B"].width = 18
    worksheet.freeze_panes = "A6"

    return worksheet


__all__ = ["build_data_quality_report"]
