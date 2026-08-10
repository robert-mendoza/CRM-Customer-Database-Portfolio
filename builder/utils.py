"""
Reusable worksheet utility functions.

Project:
    CRM Customer Database Builder

Author:
    Robert Mendoza
"""

from __future__ import annotations

from collections.abc import Sequence
from datetime import date, datetime

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from builder.styles import styles


_TITLE_ROW_HEIGHT = 24
_SUBTITLE_ROW_HEIGHT = 20
_HEADER_ROW_HEIGHT = 22

_MIN_COLUMN_WIDTH = 12
_MAX_COLUMN_WIDTH = 40
_COLUMN_PADDING = 2


__all__ = [
    "apply_headers",
    "apply_print_settings",
    "apply_title",
    "auto_fit_columns",
    "create_excel_table",
    "freeze_header",
    "set_active_cell",
]


def _apply_header_style(cell) -> None:
    """Apply the standard header style to a worksheet cell."""
    cell.font = styles.fonts.header
    cell.fill = styles.fills.header
    cell.border = styles.borders.thin
    cell.alignment = styles.alignments.center


def _calculate_text_length(value: object) -> int:
    """Return an estimated display length for a worksheet value."""
    if value is None:
        return 0

    if isinstance(value, (date, datetime)):
        return 10

    return len(str(value))


def apply_title(
    worksheet: Worksheet,
    title: str,
    subtitle: str | None = None,
) -> int:
    """Write a formatted worksheet title."""
    worksheet["A1"] = title
    worksheet["A1"].font = styles.fonts.title
    worksheet["A1"].alignment = styles.alignments.left
    worksheet.row_dimensions[1].height = _TITLE_ROW_HEIGHT

    if subtitle:
        worksheet["A2"] = subtitle
        worksheet["A2"].font = styles.fonts.body
        worksheet["A2"].alignment = styles.alignments.left
        worksheet.row_dimensions[2].height = _SUBTITLE_ROW_HEIGHT
        return 4

    return 3


def apply_headers(
    worksheet: Worksheet,
    headers: Sequence[str],
    row: int,
    enable_filter: bool = True,
) -> None:
    """Write worksheet column headers."""
    worksheet.row_dimensions[row].height = _HEADER_ROW_HEIGHT

    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(
            row=row,
            column=column_index,
            value=header,
        )
        _apply_header_style(cell)

    if enable_filter and headers:
        last_column = get_column_letter(len(headers))
        worksheet.auto_filter.ref = f"A{row}:{last_column}{row}"


def auto_fit_columns(
    worksheet: Worksheet,
    start_row: int = 1,
    end_row: int | None = None,
    min_width: int = _MIN_COLUMN_WIDTH,
    max_width: int = _MAX_COLUMN_WIDTH,
) -> None:
    """Adjust worksheet column widths based on cell content."""
    if min_width <= 0:
        raise ValueError("min_width must be greater than zero.")

    if max_width < min_width:
        raise ValueError(
            "max_width must be greater than or equal to min_width."
        )

    if start_row < 1:
        raise ValueError("start_row must be at least 1.")

    if end_row is None:
        end_row = worksheet.max_row

    if end_row < start_row:
        raise ValueError(
            "end_row must be greater than or equal to start_row."
        )

    for column_index in range(1, worksheet.max_column + 1):
        maximum_length = 0

        for row_index in range(start_row, end_row + 1):
            cell = worksheet.cell(
                row=row_index,
                column=column_index,
            )
            maximum_length = max(
                maximum_length,
                _calculate_text_length(cell.value),
            )

        if maximum_length == 0:
            continue

        width = maximum_length + _COLUMN_PADDING
        width = min(width, max_width)
        width = max(width, min_width)

        worksheet.column_dimensions[
            get_column_letter(column_index)
        ].width = width


def freeze_header(
    worksheet: Worksheet,
    header_row: int,
) -> None:
    """Freeze worksheet rows above the header row."""
    if header_row < 1:
        raise ValueError("header_row must be at least 1.")

    worksheet.freeze_panes = f"A{header_row + 1}"


def create_excel_table(
    worksheet: Worksheet,
    table_name: str,
    start_row: int,
    end_row: int,
    start_column: int = 1,
    end_column: int | None = None,
) -> Table:
    """Create and add a formatted Excel table."""
    if start_row < 1 or end_row < start_row:
        raise ValueError("Invalid table row range.")

    if start_column < 1:
        raise ValueError("start_column must be at least 1.")

    if end_column is None:
        end_column = worksheet.max_column

    if end_column < start_column:
        raise ValueError(
            "end_column must be greater than or equal to start_column."
        )

    start_cell = f"{get_column_letter(start_column)}{start_row}"
    end_cell = f"{get_column_letter(end_column)}{end_row}"

    table = Table(
        displayName=table_name,
        ref=f"{start_cell}:{end_cell}",
    )

    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    worksheet.add_table(table)
    return table


def apply_print_settings(
    worksheet: Worksheet,
    orientation: str = "landscape",
    fit_to_width: int = 1,
    repeat_header_row: int | None = None,
) -> None:
    """Apply practical print settings to a worksheet."""
    if orientation not in {"portrait", "landscape"}:
        raise ValueError(
            "orientation must be 'portrait' or 'landscape'."
        )

    if fit_to_width < 1:
        raise ValueError("fit_to_width must be at least 1.")

    if repeat_header_row is not None and repeat_header_row < 1:
        raise ValueError(
            "repeat_header_row must be at least 1."
        )

    worksheet.page_setup.orientation = orientation
    worksheet.page_setup.fitToWidth = fit_to_width
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True

    worksheet.page_margins.left = 0.25
    worksheet.page_margins.right = 0.25
    worksheet.page_margins.top = 0.5
    worksheet.page_margins.bottom = 0.5

    if repeat_header_row is not None:
        worksheet.print_title_rows = f"$1:${repeat_header_row}"


def set_active_cell(
    worksheet: Worksheet,
    cell_reference: str = "A1",
) -> None:
    """Set the active cell when the workbook opens."""
    worksheet.sheet_view.selection[0].activeCell = cell_reference
    worksheet.sheet_view.selection[0].sqref = cell_reference
