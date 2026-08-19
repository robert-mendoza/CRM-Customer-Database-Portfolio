"""Instructions worksheet builder for the CRM Customer Database Builder."""

from __future__ import annotations

from copy import copy

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from builder.constants import SheetNames
from builder.utils import apply_title


INSTRUCTION_SECTIONS: tuple[tuple[str, str], ...] = (
    ("Workbook Overview", "This workbook provides a structured CRM customer database, dashboard, data quality report, and activity log."),
    ("Customer Database", "Enter customer information in 03_Customer_Database. Use the current ten-field schema and do not overwrite the headers."),
    ("Data Validation", "Use the provided dropdown values for Industry, Customer Type, Customer Status, Lead Source, and Priority."),
    ("Activity Log", "Record customer activities in 06_Activity_Log using Activity ID, Timestamp, Customer ID, Activity Type, Description, and User. Use an existing Customer ID when recording a customer activity."),
    ("Dashboard", "Use 02_Dashboard as the summary view of the current customer database. Dashboard values depend on the customer data."),
    ("Data Quality Report", "Review 05_Data_Quality_Report after entering customer records. No Data means that no customer records currently exist."),
    ("General Guidelines", "Preserve worksheet names, headers, validation lists, and workbook structure. Review data quality after entering or updating records."),
)

CUSTOMER_DATABASE_FIELDS: tuple[str, ...] = (
    "Customer ID", "First Name", "Last Name", "Email", "Phone",
    "Industry", "Customer Type", "Customer Status", "Lead Source", "Priority",
)

ACTIVITY_LOG_FIELDS: tuple[str, ...] = (
    "Activity ID", "Timestamp", "Customer ID", "Activity Type", "Description", "User",
)


def build_instructions(workbook: Workbook) -> Worksheet:
    """Build the 07_Instructions worksheet."""
    if not isinstance(workbook, Workbook):
        raise TypeError("workbook must be an openpyxl Workbook instance.")

    worksheet_name = SheetNames.instructions
    if worksheet_name in workbook.sheetnames:
        raise ValueError(f"Worksheet '{worksheet_name}' already exists.")

    worksheet = workbook.create_sheet(title=worksheet_name)
    apply_title(worksheet, "CRM Workbook Instructions")

    row = 4
    for section, description in INSTRUCTION_SECTIONS:
        worksheet.cell(row=row, column=1, value=section)
        worksheet.cell(row=row, column=2, value=description)
        row += 1

    worksheet.cell(row=row + 1, column=1, value="Customer Database Fields")
    row += 2
    for field in CUSTOMER_DATABASE_FIELDS:
        worksheet.cell(row=row, column=1, value=field)
        row += 1

    row += 1
    worksheet.cell(row=row, column=1, value="Activity Log Fields")
    row += 1
    for field in ACTIVITY_LOG_FIELDS:
        worksheet.cell(row=row, column=1, value=field)
        row += 1

    worksheet.column_dimensions["A"].width = 30
    worksheet.column_dimensions["B"].width = 100
    worksheet.sheet_view.zoomScale = 100
    worksheet.freeze_panes = "A4"

    for cell in worksheet["B"]:
        alignment = copy(cell.alignment)
        alignment.wrap_text = True
        alignment.vertical = "top"
        cell.alignment = alignment

    return worksheet
