"""Validation lists worksheet builder.

Creates the worksheet containing controlled values used by
customer database Excel data-validation lists.

Project:
    CRM Customer Database Builder

Author:
    Robert Mendoza
"""

from __future__ import annotations

from collections.abc import Sequence

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from builder.constants import SheetNames
from builder.validation_lists import (
    CUSTOMER_STATUSES,
    CUSTOMER_TYPES,
    INDUSTRIES,
    LEAD_SOURCES,
    PRIORITIES,
)


VALIDATION_LIST_HEADERS: tuple[str, ...] = (
    "Customer Statuses",
    "Customer Types",
    "Industries",
    "Lead Sources",
    "Priorities",
)

VALIDATION_LISTS: tuple[Sequence[str], ...] = (
    CUSTOMER_STATUSES,
    CUSTOMER_TYPES,
    INDUSTRIES,
    LEAD_SOURCES,
    PRIORITIES,
)


def build_validation_lists(workbook: Workbook) -> Worksheet:
    """Create the validation lists worksheet.

    Args:
        workbook: Target workbook.

    Returns:
        The validation lists worksheet.
    """
    if not isinstance(workbook, Workbook):
        raise TypeError("workbook must be an openpyxl Workbook instance.")

    worksheet_name = SheetNames.validation_lists

    if worksheet_name in workbook.sheetnames:
        worksheet = workbook[worksheet_name]
    else:
        worksheet = workbook.create_sheet(worksheet_name)

    for column_index, header in enumerate(
        VALIDATION_LIST_HEADERS,
        start=1,
    ):
        worksheet.cell(
            row=1,
            column=column_index,
            value=header,
        )

        values = VALIDATION_LISTS[column_index - 1]

        for row_index, value in enumerate(values, start=2):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

    return worksheet


__all__ = [
    "VALIDATION_LIST_HEADERS",
    "VALIDATION_LISTS",
    "build_validation_lists",
]