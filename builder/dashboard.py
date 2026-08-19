"""Dashboard worksheet builder for the CRM Customer Database Builder.

Project:
    CRM Customer Database Builder

Author:
    Robert Mendoza
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from builder.constants import SHEETS, WORKBOOK
from builder.styles import styles
from builder.utils import apply_title

_CUSTOMER_ID_COLUMN = "A"
_CUSTOMER_STATUS_COLUMN = "H"
_CUSTOMER_DATA_START_ROW = 5
_CUSTOMER_DATA_END_ROW = 1004

_KPI_ROWS: tuple[tuple[str, int], ...] = (
    ("Total Customers", 5),
    ("Active Customers", 6),
    ("Prospect Customers", 7),
    ("Inactive Customers", 8),
)


def build_dashboard(
    workbook: Workbook,
) -> Worksheet:
    """Create the dashboard worksheet with dynamic customer KPIs.

    KPI formulas reference the customer database worksheet so the dashboard
    updates when customer records are added or their status changes in Excel.

    Args:
        workbook: Target openpyxl workbook.

    Returns:
        The created dashboard worksheet.

    Raises:
        TypeError: If ``workbook`` is not an openpyxl Workbook instance.
        ValueError: If the customer database worksheet does not exist.
    """
    if not isinstance(workbook, Workbook):
        raise TypeError("workbook must be an openpyxl Workbook instance.")

    if SHEETS.customer_database not in workbook.sheetnames:
        raise ValueError(
            f"Required worksheet '{SHEETS.customer_database}' does not exist."
        )

    worksheet = workbook.create_sheet(title=SHEETS.dashboard)

    apply_title(
        worksheet,
        "CRM Dashboard",
        WORKBOOK.company,
    )

    worksheet["A4"] = "KPI"
    worksheet["B4"] = "Value"

    for cell in worksheet[4]:
        if cell.column <= 2:
            cell.font = styles.fonts.header
            cell.fill = styles.fills.header
            cell.border = styles.borders.thin
            cell.alignment = styles.alignments.center

    data_range = (
        f"'{SHEETS.customer_database}'!"
        f"{_CUSTOMER_ID_COLUMN}{_CUSTOMER_DATA_START_ROW}:"
        f"{_CUSTOMER_ID_COLUMN}{_CUSTOMER_DATA_END_ROW}"
    )
    status_range = (
        f"'{SHEETS.customer_database}'!"
        f"{_CUSTOMER_STATUS_COLUMN}{_CUSTOMER_DATA_START_ROW}:"
        f"{_CUSTOMER_STATUS_COLUMN}{_CUSTOMER_DATA_END_ROW}"
    )

    formulas = (
        f"=COUNTA({data_range})",
        f'=COUNTIF({status_range},"Active")',
        f'=COUNTIF({status_range},"Prospect")',
        f'=COUNTIF({status_range},"Inactive")',
    )

    for (label, row), formula in zip(_KPI_ROWS, formulas):
        worksheet.cell(row=row, column=1, value=label)
        worksheet.cell(row=row, column=2, value=formula)

        worksheet.cell(row=row, column=1).font = styles.fonts.body
        worksheet.cell(row=row, column=2).font = styles.fonts.body

        for column in range(1, 3):
            cell = worksheet.cell(row=row, column=column)
            cell.border = styles.borders.thin
            cell.alignment = styles.alignments.left

    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions["B"].width = 18
    worksheet.sheet_view.zoomScale = 100

    return worksheet


__all__ = ["build_dashboard"]
