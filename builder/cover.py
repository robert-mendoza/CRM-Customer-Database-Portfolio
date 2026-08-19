"""Cover worksheet builder for the CRM Customer Database Builder.

Project:
    CRM Customer Database Builder

Author:
    Robert Mendoza
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from builder.constants import SheetNames, WorkbookConfig
from builder.styles import styles
from builder.utils import apply_title


def build_cover(
    workbook: Workbook,
    config: WorkbookConfig,
) -> Worksheet:
    """Create the workbook cover worksheet.

    Args:
        workbook: Target openpyxl workbook.
        config: Workbook metadata used for the cover content.

    Returns:
        The created cover worksheet.

    Raises:
        TypeError: If ``workbook`` or ``config`` has an invalid type.
    """
    if not isinstance(workbook, Workbook):
        raise TypeError("workbook must be an openpyxl Workbook instance.")
    if not isinstance(config, WorkbookConfig):
        raise TypeError("config must be a WorkbookConfig instance.")

    worksheet = workbook.create_sheet(title=SheetNames.cover)

    apply_title(
        worksheet,
        config.title,
        config.company,
    )

    worksheet["A4"] = "Version"
    worksheet["B4"] = config.version
    worksheet["A5"] = "Author"
    worksheet["B5"] = config.author
    worksheet["A6"] = "File Name"
    worksheet["B6"] = config.filename

    for row in range(4, 7):
        worksheet.cell(row=row, column=1).font = styles.fonts.header
        worksheet.cell(row=row, column=1).fill = styles.fills.header
        worksheet.cell(row=row, column=1).border = styles.borders.thin
        worksheet.cell(row=row, column=1).alignment = styles.alignments.left
        worksheet.cell(row=row, column=2).font = styles.fonts.body
        worksheet.cell(row=row, column=2).border = styles.borders.thin
        worksheet.cell(row=row, column=2).alignment = styles.alignments.left

    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 48
    worksheet.sheet_view.zoomScale = 100

    return worksheet


__all__ = ["build_cover"]
