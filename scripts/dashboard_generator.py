"""
Dashboard Generator

CRM Customer Database Portfolio

Version: 1.3
"""

from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill


HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True,
    size=16,
)

SECTION_FONT = Font(
    bold=True,
    size=11,
)

VALUE_FONT = Font(
    bold=True,
    size=12,
)

CENTER = Alignment(
    horizontal="center",
    vertical="center",
)


def create_dashboard(workbook):
    """
    Create the Sales Dashboard worksheet.
    """

    if "Sales Dashboard" in workbook.sheetnames:
        del workbook["Sales Dashboard"]

    ws = workbook.create_sheet(
        title="Sales Dashboard",
        index=0,
    )

    #
    # Dashboard Title
    #

    ws.merge_cells("A1:B1")

    ws["A1"] = "CRM SALES DASHBOARD"

    ws["A1"].fill = HEADER_FILL
    ws["A1"].font = HEADER_FONT
    ws["A1"].alignment = CENTER

    #
    # KPI Section
    #

    ws["A3"] = "Total Customers"
    ws["A4"] = "Active Customers"
    ws["A5"] = "Prospect Customers"
    ws["A6"] = "Inactive Customers"
    ws["A7"] = "Total Revenue"
    ws["A8"] = "Average Contract Value"
    ws["A9"] = "Overdue Payments"

    ws["B3"] = "=COUNTA(Customers!A:A)-1"
    ws["B4"] = '=COUNTIF(Customers!K:K,"Active")'
    ws["B5"] = '=COUNTIF(Customers!K:K,"Prospect")'
    ws["B6"] = '=COUNTIF(Customers!K:K,"Inactive")'
    ws["B7"] = "=SUM(Customers!Q:Q)"
    ws["B8"] = "=AVERAGE(Customers!Q:Q)"
    ws["B9"] = '=COUNTIF(Customers!R:R,"Overdue")'

    #
    # Country Summary
    #

    ws["D3"] = "Country"
    ws["E3"] = "Customers"

    countries = [
        "United States",
        "Canada",
        "United Kingdom",
        "Australia",
        "Singapore",
        "Philippines",
        "Germany",
        "Japan",
        "Netherlands",
        "New Zealand",
    ]

    row = 4

    for country in countries:

        ws[f"D{row}"] = country

        ws[f"E{row}"] = (
            f'=COUNTIF(Customers!I:I,"{country}")'
        )

        row += 1

    #
    # Industry Summary
    #

    ws["G3"] = "Industry"
    ws["H3"] = "Customers"

    industries = [
        "Software",
        "Finance",
        "Healthcare",
        "Manufacturing",
        "Retail",
        "Logistics",
        "Telecommunications",
        "Education",
        "Consulting",
    ]

    row = 4

    for industry in industries:

        ws[f"G{row}"] = industry

        ws[f"H{row}"] = (
            f'=COUNTIF(Customers!F:F,"{industry}")'
        )

        row += 1

    #
    # Revenue by Account Manager
    #

    ws["J3"] = "Account Manager"
    ws["K3"] = "Revenue"

    managers = [
        "Robert Mendoza",
        "Anna Cruz",
        "John Reyes",
        "Maria Santos",
        "David Lim",
        "Jennifer Lee",
        "Michael Garcia",
        "Sarah Johnson",
    ]

    row = 4

    for manager in managers:

        ws[f"J{row}"] = manager

        ws[f"K{row}"] = (
            f'=SUMIF(Customers!M:M,J{row},Customers!Q:Q)'
        )

        row += 1

    #
    # Formatting
    #

    for cell in [
        "A3",
        "A4",
        "A5",
        "A6",
        "A7",
        "A8",
        "A9",
        "D3",
        "E3",
        "G3",
        "H3",
        "J3",
        "K3",
    ]:

        ws[cell].font = SECTION_FONT

    for cell in [
        "B3",
        "B4",
        "B5",
        "B6",
        "B7",
        "B8",
        "B9",
    ]:

        ws[cell].font = VALUE_FONT

    ws["B7"].number_format = '"$"#,##0.00'
    ws["B8"].number_format = '"$"#,##0.00'

    ws.freeze_panes = "A2"

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["J"].width = 24
    ws.column_dimensions["K"].width = 18

    return ws