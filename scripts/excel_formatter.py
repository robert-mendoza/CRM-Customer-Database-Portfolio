"""
Excel Formatting Utilities
CRM Customer Database Portfolio
"""

from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import Alignment

from openpyxl.worksheet.table import Table
from openpyxl.worksheet.table import TableStyleInfo

from openpyxl.utils import get_column_letter

from config import (
    HEADER_BACKGROUND,
    HEADER_FONT,
    TABLE_NAME,
    TABLE_STYLE,
    FREEZE_PANE,
    DATE_FORMAT,
    CURRENCY_FORMAT,
)

# ------------------------------------------------------------------
# Styles
# ------------------------------------------------------------------

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor=HEADER_BACKGROUND,
)

HEADER_FONT_STYLE = Font(
    bold=True,
    color=HEADER_FONT,
)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

CENTER = Alignment(
    horizontal="center",
    vertical="center",
)

LEFT = Alignment(
    horizontal="left",
    vertical="center",
)

# ------------------------------------------------------------------
# Header Formatting
# ------------------------------------------------------------------

def format_header(ws):

    for cell in ws[1]:

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT_STYLE
        cell.border = THIN_BORDER
        cell.alignment = CENTER

# ------------------------------------------------------------------
# Freeze Top Row
# ------------------------------------------------------------------

def freeze_top_row(ws):

    ws.freeze_panes = FREEZE_PANE

# ------------------------------------------------------------------
# Auto Filter
# ------------------------------------------------------------------

def enable_auto_filter(ws):

    ws.auto_filter.ref = ws.dimensions

# ------------------------------------------------------------------
# AutoFit Columns
# ------------------------------------------------------------------

def autofit_columns(ws):

    for column_cells in ws.columns:

        max_length = 0

        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:

            try:

                length = len(str(cell.value))

                if length > max_length:
                    max_length = length

            except Exception:
                pass

        ws.column_dimensions[column_letter].width = min(max_length + 3, 40)

# ------------------------------------------------------------------
# Currency Format
# ------------------------------------------------------------------

def apply_currency_format(ws, column_letter):

    for row in range(2, ws.max_row + 1):

        ws[f"{column_letter}{row}"].number_format = CURRENCY_FORMAT

# ------------------------------------------------------------------
# Date Format
# ------------------------------------------------------------------

def apply_date_format(ws, columns):

    for column in columns:

        for row in range(2, ws.max_row + 1):

            ws[f"{column}{row}"].number_format = DATE_FORMAT

# ------------------------------------------------------------------
# Borders
# ------------------------------------------------------------------

def apply_borders(ws):

    for row in ws.iter_rows():

        for cell in row:

            cell.border = THIN_BORDER

            if cell.row > 1:

                cell.alignment = LEFT

# ------------------------------------------------------------------
# Excel Table
# ------------------------------------------------------------------

def create_excel_table(ws):

    table = Table(

        displayName=TABLE_NAME,

        ref=ws.dimensions,

    )

    style = TableStyleInfo(

        name=TABLE_STYLE,

        showFirstColumn=False,

        showLastColumn=False,

        showRowStripes=True,

        showColumnStripes=False,

    )

    table.tableStyleInfo = style

    ws.add_table(table)

# ------------------------------------------------------------------
# Apply All Formatting
# ------------------------------------------------------------------

def apply_all_formatting(ws):

    format_header(ws)

    freeze_top_row(ws)

    enable_auto_filter(ws)

    autofit_columns(ws)

    apply_borders(ws)

    create_excel_table(ws)

    #
    # Contract Value
    #
    apply_currency_format(ws, "Q")

    #
    # Dates
    #
    apply_date_format(
        ws,
        ["N", "O", "P"],
    )